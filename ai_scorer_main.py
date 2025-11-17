import os
import json
import time
import pandas as pd
from tqdm import tqdm
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.runnables import RunnableSequence

#  CONFIG
load_dotenv()
API_KEY = os.getenv("GOOGLE_API_KEY")

INPUT_FILE = "sample_50.xlsx"
OUTPUT_FILE = "radiology_llm_output.xlsx"   # single Excel file (2 sheets)
JSON_FILE = "config.json"

MODEL_NAME = "gemini-2.0-flash-001"
DELAY_BETWEEN_REPORTS = 2
MAX_RETRIES = 6

# Classifier names for confusion matrix
CLASSIFIERS = [
    "pneumonia",
    "pulmonary_nodules",
    "bronchitis",
    "rtm",
    "focal_caudodorsal_lung",
    "interstitial",
    "diseased_lungs",
    "perihilar_infiltrate",
    "hypo_plastic_trachea",
    "cardiomegaly",
    "pleural_effusion",
    "focal_perihilar",
    "pulmonary_hypoinflation",
    "right_sided_cardiomegaly",
    "pericardial_effusion",
    "bronchiectasis",
    "pulmonary_vessel_enlargement",
    "left_sided_cardiomegaly",
    "thoracic_lymphadenopathy",
    "esophagitis",
    "vhs_v2",
]

# Output column structure
OUTPUT_COLUMNS = [
    "condition", "true_Positive", "false_Negative",
    "true_Negative", "false_Positive", "Sensitivity",
    "Specificity", "Check", "Positive Ground Truth",
    "Negative Ground Truth", "Ground Truth Check"
]

# ---------------- LOAD CONDITIONS ----------------
with open(JSON_FILE, "r", encoding="utf-8") as f:
    conditions_list = json.load(f)
CONDITIONS = [c["condition"] for c in conditions_list]

print(f" Loaded {len(CONDITIONS)} conditions from {JSON_FILE}")
print(f"Example: {', '.join(CONDITIONS[:5])}... (+{len(CONDITIONS)-5} more)\n")

# ---------------- LOAD DATA ----------------
df = pd.read_excel(INPUT_FILE)
required_cols = [
    "Findings (original radiologist report)",
    "Conclusions (original radiologist report)",
    "Findings (AI report)",
    "Conclusions (AI report)",
]
for col in required_cols:
    if col not in df.columns:
        raise ValueError(f"Missing column: {col}")

# MODEL INIT
llm = ChatGoogleGenerativeAI(
    model=MODEL_NAME,
    temperature=0,
    google_api_key=API_KEY
)

# PROMPT TEMPLATE
batch_prompt_template = ChatPromptTemplate.from_template("""
You are an expert veterinary radiologist.

Evaluate the following report for the presence of these conditions:

{conditions_with_descriptions}

For each condition, return strictly:
condition_name: Positive or Negative

Report Findings:
{findings}

Report Conclusion:
{conclusions}
""")

batch_chain = RunnableSequence(batch_prompt_template | llm)


# ---------------- DETECTION FUNCTION ----------------

def _detect_normalize(name: str) -> str:
    s = name.lower().replace("_", " ").strip()
    return "".join(ch for ch in s if ch.isalnum())


def _detect_build_normalized_map(conds):
    nm = {}
    for c in conds:
        norm = _detect_normalize(c)
        nm[norm] = c
        nm[c.lower().replace("_", " ")] = c
    return nm


def _detect_parse_response_text(text: str, normalized_map: dict) -> dict:
    result = {}
    lines = [ln.strip() for ln in text.splitlines() if ":" in ln]
    for line in lines:
        left, right = line.split(":", 1)
        key = _detect_normalize(left)
        val = right.strip()
        match = normalized_map.get(key)
        if not match:
            for k, v in normalized_map.items():
                if k and (k in key or key in k):
                    match = v
                    break
        if match:
            result[match] = val
    for c in CONDITIONS:
        result.setdefault(c, "Negative")
    return result


def _detect_invoke_with_retries(payload: dict, retries: int) -> str:
    for attempt in range(retries):
        try:
            response = batch_chain.invoke(payload)
            return (
                response.content.strip()
                if hasattr(response, "content")
                else str(response).strip()
            )
        except Exception as e:
            msg = str(e).lower()
            if "429" in msg or "quota" in msg:
                print(f" Quota limit hit Waiting 60s (attempt {attempt+1})...")
                time.sleep(60)
                continue
            print(f" Error: {e}")
            return ""
    return ""


def detect_all_conditions(findings, conclusions, max_retries=MAX_RETRIES):
    conditions_with_descriptions = "\n".join(
        [
            f"- {c['condition']}: {c['prompt']}"
            for c in conditions_list
        ]
    )
    normalized_map = _detect_build_normalized_map(CONDITIONS)

    payload = {
        "conditions_with_descriptions": conditions_with_descriptions,
        "findings": str(findings),
        "conclusions": str(conclusions)
    }
    text = _detect_invoke_with_retries(payload, max_retries)
    if not text:
        return dict.fromkeys(CONDITIONS, "Negative")
    return _detect_parse_response_text(text, normalized_map)


# ---------------- MAIN LLM LOOP ----------------
progress = tqdm(total=len(df), desc="LLM Evaluating", ncols=100)
for idx, row in df.iterrows():
    if all(
        f"{c}_original" in df.columns
        and pd.notna(df.at[idx, f"{c}_original"])
        for c in CONDITIONS
    ):
        progress.update(1)
        continue

    results_orig = detect_all_conditions(
        row["Findings (original radiologist report)"],
        row["Conclusions (original radiologist report)"]
    )
    results_ai = detect_all_conditions(
        row["Findings (AI report)"],
        row["Conclusions (AI report)"]
    )
    for c in CONDITIONS:
        df.at[idx, f"{c}_original"] = results_orig[c]
        df.at[idx, f"{c}_ai"] = results_ai[c]
    if (idx + 1) % 5 == 0:
        temp_file = OUTPUT_FILE.replace(".xlsx", "_partial.xlsx")
        df.to_excel(temp_file, index=False)
        print(f"\n Autosaved after {idx + 1} rows")
    time.sleep(DELAY_BETWEEN_REPORTS)
    progress.update(1)
progress.close()
print("\n LLM predictions completed.")


# ---------------- LABEL TP / FP / TN / FN ----------------
def label_predictions(df):
    conditions = [
        col.replace("_original", "")
        for col in df.columns
        if col.endswith("_original")
    ]
    df["true_positive"] = ""
    df["true_negative"] = ""
    df["false_positive"] = ""
    df["false_negative"] = ""
    for i, row in df.iterrows():
        tp, tn, fp, fn = [], [], [], []
        for cond in conditions:
            orig = str(row.get(f"{cond}_original", "")).strip().lower()
            ai = str(row.get(f"{cond}_ai", "")).strip().lower()
            if orig == "positive" and ai == "positive":
                tp.append(cond)
            elif orig == "negative" and ai == "negative":
                tn.append(cond)
            elif orig == "negative" and ai == "positive":
                fp.append(cond)
            elif orig == "positive" and ai == "negative":
                fn.append(cond)
        df.at[i, "true_positive"] = ", ".join(tp)
        df.at[i, "true_negative"] = ", ".join(tn)
        df.at[i, "false_positive"] = ", ".join(fp)
        df.at[i, "false_negative"] = ", ".join(fn)
    return df


# ---------------- UPDATED CONFUSION MATRIX BUILDER (ORDER FIXED) ----------------

def build_confusion(df):

    print("\nBuilding Confusion Matrix with UPDATED COLUMN...\n")

    # Normalize
    norm_df = df.copy()
    norm_df.columns = (
        norm_df.columns.str.strip().str.lower().str.replace(" ", "_")
    )

    # Detect actual column names
    col_map = {
        "true_positive": next((c for c in norm_df.columns if "true_positive" in c), None),
        "false_positive": next((c for c in norm_df.columns if "false_positive" in c), None),
        "true_negative": next((c for c in norm_df.columns if "true_negative" in c), None),
        "false_negative": next((c for c in norm_df.columns if "false_negative" in c), None),
    }

    for key, col in col_map.items():
        if col is None:
            print(f"⚠ Missing {key}. Creating blank column.")
            norm_df[key] = ""
        else:
            norm_df[key] = norm_df[col].fillna("").astype(str)

    result_rows = []

    print("Condition Counts (TP, FN, TN, FP) ")
    print("------------------------------------------------")

    for cond in CLASSIFIERS:

        tp = norm_df["true_positive"].str.contains(fr"\b{cond}\b", case=False, na=False, regex=True).sum()
        fn = norm_df["false_negative"].str.contains(fr"\b{cond}\b", case=False, na=False, regex=True).sum()
        tn = norm_df["true_negative"].str.contains(fr"\b{cond}\b", case=False, na=False, regex=True).sum()
        fp = norm_df["false_positive"].str.contains(fr"\b{cond}\b", case=False, na=False, regex=True).sum()

        print(f"{cond:30} | TP:{tp}  FN:{fn}  TN:{tn}  FP:{fp}")

        # NOTE: ORDER CHANGED EXACTLY AS YOU REQUESTED:
        # B = TP
        # C = FN
        # D = TN
        # E = FP

        result_rows.append([
            cond,   # A
            tp,     # truePositive
            fn,     # falseNegative
            tn,     # trueNegative
            fp,     # falsePositive
            "", "", "", "", "", ""
        ])

    cm_df = pd.DataFrame(result_rows, columns=OUTPUT_COLUMNS)

    # ---------------- WRITE EXCEL ----------------

    wb = Workbook()
    ws = wb.active
    ws.title = "Confusion Matrix"

    for r in dataframe_to_rows(cm_df, index=False, header=True):
        ws.append(r)

    # UPDATED FORMULAS → MATCH NEW COLUMN ORDER
    for i in range(2, len(cm_df) + 2):

        # Sensitivity = TP / (TP + FN)
        ws[f"F{i}"] = f"=IFERROR(B{i}/(B{i}+C{i}),0)"

        # Specificity = TN / (TN + FP)
        ws[f"G{i}"] = f"=IFERROR(D{i}/(D{i}+E{i}),0)"

        # Check = B + C + D + E
        ws[f"H{i}"] = f"=SUM(B{i},C{i},D{i},E{i})"

        # Positive Ground Truth = TP + FN
        ws[f"I{i}"] = f"=SUM(B{i},C{i})"

        # Negative Ground Truth = TN + FP
        ws[f"J{i}"] = f"=SUM(D{i},E{i})"

        # Ground Truth Check = I + J
        ws[f"K{i}"] = f"=SUM(I{i},J{i})"

        ws[f"F{i}"].number_format = "0.00%"
        ws[f"G{i}"].number_format = "0.00%"

    # Add raw predictions sheet
    ws_input = wb.create_sheet(title="LLM Predictions")
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_input.append(r)

    # Formatting
    for sheet in [ws, ws_input]:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(OUTPUT_FILE)
    print(f"\nConfusion Matrix updated with new TP/FN/TN/FP order — saved to {OUTPUT_FILE}\n")


# FINAL EXECUTION
df = label_predictions(df)
build_confusion(df)
print("\nAll done — single Excel file with Confusion Matrix & Predictions generated successfully!")	 