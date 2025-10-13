import sys
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

# Classifiers
CLASSIFIERS = [
    "pneumonia",
    "pulmonary_nodules",
    "bronchitis",
    "rtm",
    "focal_caudodorsal_lung",
    "interstitial",
    "diseased_lungs",
    "perihilar_infiltrate",
    "hypo_plastic_trachea",
    "cardiomegaly",
    "pleural_effusion",
    "focal_perihilar",
    "pulmonary_hypoinflation",
    "right_sided_cardiomegaly",
    "pericardial_effusion",
    "bronchiectasis",
    "pulmonary_vessel_enlargement",
    "left_sided_cardiomegaly",
    "thoracic_lymphadenopathy",
    "esophagitis",
    "vhs_v2",
]

# Output columns
OUTPUT_COLUMNS = [
    "condition",
    "true_Positive",
    "false_Positive",
    "true_Negative",
    "false_Negative",
    "Sensitivity",
    "Specificity",
    "Check",
    "Positive Ground Truth",
    "Negative Ground Truth",
    "Ground Truth Check",
]


def read_excel(input_path: Path) -> pd.DataFrame:
    """Read Excel input file safely."""
    try:
        df = pd.read_excel(input_path)
        print(f"Successfully read input: {input_path.name}")
        return df
    except Exception as e:
        print(f"Error reading file: {e}")
        sys.exit(1)


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize column names (lowercase and strip)."""
    df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")
    return df


def count_classifiers(df: pd.DataFrame) -> pd.DataFrame:
    """Count each classifier occurrence in confusion matrix columns."""
    df = normalize_columns(df)

    # Map columns safely
    col_map = {
        "true_positive": next(
            (
                c
                for c in df.columns
                if "true_positive" in c
            ),
            None,
        ),
        "false_positive": next(
            (
                c
                for c in df.columns
                if "false_positive" in c
            ),
            None
        ),
        "true_negative": next(
            (
                c
                for c in df.columns
                if "true_negative" in c
            ),
            None,
        ),
        "false_negative": next(
            (c for c in df.columns if "false_negative" in c),
            None,
        ),
    }

    for k, v in col_map.items():
        if v is None:
            print(f"Warning: Column '{k}' not found in Excel.")
            df[k] = ""
        else:
            df[k] = df[v].fillna("").astype(str)

    result_rows = []

    for cond in CLASSIFIERS:
        tp = df["true_positive"].str.contains(
            cond, case=False, na=False).sum()
        fp = df["false_positive"].str.contains(
            cond, case=False, na=False).sum()
        tn = df["true_negative"].str.contains(
            cond, case=False, na=False).sum()
        fn = df["false_negative"].str.contains(
            cond, case=False, na=False).sum()

        result_rows.append(
            [cond, tp, fp, tn, fn, "", "", "", "", "", ""]
        )

    result_df = pd.DataFrame(result_rows, columns=OUTPUT_COLUMNS)
    return result_df


def write_excel_with_formulas(
    df: pd.DataFrame, input_df: pd.DataFrame, output_path: Path
):
    """Write the result DataFrame to Excel with formulas"""
    wb = Workbook()

    # First sheet: Confusion Matrix results
    ws = wb.active
    ws.title = "Confusion Matrix"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Add formulas for each row
    for i in range(2, len(df) + 2):
        # Sensitivity formula: =IFERROR(SUM(B2:B2)/(SUM(B2:B2)+SUM(E2:E2)), 0)
        ws[f"F{i}"] = (
            f"=IFERROR(SUM(B{i}:B{i})/"
            f"(SUM(B{i}:B{i})+SUM(E{i}:E{i})), 0)"
        )
        # Specificity formula
        ws[f"G{i}"] = f"=IFERROR(D{i}/(D{i}+C{i}),0)"
        # Format Sensitivity and Specificity columns as percentage
        ws[f"F{i}"].number_format = '0.00%'
        ws[f"G{i}"].number_format = '0.00%'

        # Check column formula: =SUM(true_Positive+false_Positive)
        ws[f"H{i}"] = f"=SUM(B{i}+C{i})"

        # Positive Ground Truth formula: =SUM(B2:E2)
        ws[f"I{i}"] = f"=SUM(B{i}:E{i})"

        # Negative Ground Truth formula: =SUM(D2:C2)
        ws[f"J{i}"] = f"=SUM(D{i}:C{i})"

        # Ground Truth Check formula: =SUM(I2:J2)
        ws[f"K{i}"] = f"=SUM(I{i}:J{i})"

    # Second sheet: Original input data
    ws_input = wb.create_sheet(title="Input Data")

    for r in dataframe_to_rows(input_df, index=False, header=True):
        ws_input.append(r)

    # Center-align cells in both sheets
    for ws_sheet in [ws_input, ws]:
        for row in ws_sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center")

    wb.save(output_path)
    print(f"Output file created: {output_path}")


def main():
    if len(sys.argv) < 2:
        print("Usage: python script.py <input_excel_path>")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    if not input_path.exists():
        print(f"File not found: {input_path}")
        sys.exit(1)

    df = read_excel(input_path)
    output_df = count_classifiers(df)

    output_path = input_path.parent / "output_confusion_matrix.xlsx"
    write_excel_with_formulas(output_df, df, output_path)


if __name__ == "__main__":
    main()
