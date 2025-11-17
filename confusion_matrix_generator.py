import sys
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

# Classifiers
CLASSIFIERS = [
    "pneumonia",
    "pulmonary_nodules",
    "bronchitis",
    "rtm",
    "focal_caudodorsal_lung",
    "interstitial",
    "diseased_lungs",
    "perihilar_infiltrate",
    "hypo_plastic_trachea",
    "cardiomegaly",
    "pleural_effusion",
    "focal_perihilar",
    "pulmonary_hypoinflation",
    "right_sided_cardiomegaly",
    "pericardial_effusion",
    "bronchiectasis",
    "pulmonary_vessel_enlargement",
    "left_sided_cardiomegaly",
    "thoracic_lymphadenopathy",
    "esophagitis",
    "vhs_v2",
]

OUTPUT_COLUMNS = [
    "condition",
    "true_Positive",
    "false_Negative",
    "true_Negative",
    "false_Positive",
    "Sensitivity",
    "Specificity",
    "Check",
    "Positive Ground Truth",
    "Negative Ground Truth",
    "Ground Truth Check",
]


def read_excel(input_path: Path) -> pd.DataFrame:
    print("Reading Excel File...")
    try:
        df = pd.read_excel(input_path)
        print(f"Successfully read input file: {input_path.name}")
        print(f"Total Rows Loaded: {len(df)}")
        return df
    except Exception as e:
        print(f"Error reading file: {e}")
        sys.exit(1)


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    print("Normalizing column names...")
    df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")
    print("Column normalization complete")
    print(f"Columns Found: {list(df.columns)}")
    return df


def count_classifiers(df: pd.DataFrame) -> pd.DataFrame:
    print("Counting classifiers...")
    df = normalize_columns(df)

    col_map = {
        "true_positive": next((c for c in df.columns if "true_positive" in c), None),
        "false_positive": next((c for c in df.columns if "false_positive" in c), None),
        "true_negative": next((c for c in df.columns if "true_negative" in c), None),
        "false_negative": next((c for c in df.columns if "false_negative" in c), None),
    }

    print("Mapped Columns:")
    for k, v in col_map.items():
        print(f"  {k}: {v}")

    for k, v in col_map.items():
        if v is None:
            print(f"Warning: Column '{k}' not found! Creating empty field.")
            df[k] = ""
        else:
            df[k] = df[v].fillna("").astype(str)

    result_rows = []

    print("\nPer-Condition Confusion Matrix Counts:")
    print("------------------------------------------")

    for cond in CLASSIFIERS:
        tp = df["true_positive"].str.contains(fr'\b{cond}\b', case=False, na=False, regex=True).sum()
        fp = df["false_positive"].str.contains(fr'\b{cond}\b', case=False, na=False, regex=True).sum()
        tn = df["true_negative"].str.contains(fr'\b{cond}\b', case=False, na=False, regex=True).sum()
        fn = df["false_negative"].str.contains(fr'\b{cond}\b', case=False, na=False, regex=True).sum()

        print(f"{cond:30} | TP: {tp}  FP: {fp}  TN: {tn}  FN: {fn}")

        result_rows.append([cond, tp,  fn, tn, fp, "", "", "", "", "", ""])

    print("\nClassifier counting completed.")
    return pd.DataFrame(result_rows, columns=OUTPUT_COLUMNS)


def write_excel_with_formulas(df: pd.DataFrame, input_df: pd.DataFrame, output_path: Path):
    print("Writing Excel Output File...")
    wb = Workbook()

    ws = wb.active
    ws.title = "Confusion Matrix"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    for i in range(2, len(df) + 2):

        ws[f"F{i}"] = f"=IFERROR(B{i}/(B{i}+C{i}),0)"
        ws[f"G{i}"] = f"=IFERROR(D{i}/(D{i}+E{i}),0)"
        ws[f"H{i}"] = f"=SUM(B{i},C{i},D{i},E{i})"
        ws[f"I{i}"] = f"=SUM(B{i},C{i})"
        ws[f"J{i}"] = f"=SUM(D{i},E{i})"
        ws[f"K{i}"] = f"=SUM(I{i},J{i})"

        ws[f"F{i}"].number_format = "0.00%"
        ws[f"G{i}"].number_format = "0.00%"

    ws_input = wb.create_sheet(title="Input Data")

    for r in dataframe_to_rows(input_df, index=False, header=True):
        ws_input.append(r)

    print("Applying cell alignment...")

    for ws_sheet in [ws_input, ws]:
        for row in ws_sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(output_path)
    print(f"Excel file created: {output_path}")


def main():
    print("Script Started")

    if len(sys.argv) < 2:
        print("Usage: python script.py <input_excel_path>")
        sys.exit(1)

    input_path = Path(sys.argv[1])

    if not input_path.exists():
        print(f"File not found: {input_path}")
        sys.exit(1)

    df = read_excel(input_path)
    output_df = count_classifiers(df)

    output_path = input_path.parent / "output_confusion_matrix.xlsx"
    write_excel_with_formulas(output_df, df, output_path)

    print("Script Finished Successfully!")


if __name__ == "__main__":
    main()
