import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

# SETTINGS
INPUT_FILE = "radiology_reports_with_predictions_all_conditions.xlsx"
LABEL_OUTPUT = "llm_label_output_with_results.xlsx"
CONFUSION_OUTPUT = "output_confusion_matrix.xlsx"

# Master list of classifier names
CLASSIFIERS = [
    "pneumonia",
    "pulmonary_nodules",
    "bronchitis",
    "rtm",
    "focal_caudodorsal_lung",
    "interstitial",
    "diseased_lungs",
    "perihilar_infiltrate",
    "hypo_plastic_trachea",
    "cardiomegaly",
    "pleural_effusion",
    "focal_perihilar",
    "pulmonary_hypoinflation",
    "right_sided_cardiomegaly",
    "pericardial_effusion",
    "bronchiectasis",
    "pulmonary_vessel_enlargement",
    "left_sided_cardiomegaly",
    "thoracic_lymphadenopathy",
    "esophagitis",
    "vhs_v2",
]

OUTPUT_COLUMNS = [
    "condition", "true_Positive", "false_Positive",
    "true_Negative", "false_Negative", "Sensitivity",
    "Specificity", "Check", "Positive Ground Truth",
    "Negative Ground Truth", "Ground Truth Check"
]


#  STEP 1 — LABEL TP/FP/TN/FN
# DETERMIN CONDITION TO CHECK
def label_predictions(df):
    conditions = [
        col.replace("_original", "")
        for col in df.columns
        if col.endswith("_original")
    ]
# PREPARE COLUMNS
    df["true_positive"] = ""
    df["true_negative"] = ""
    df["false_positive"] = ""
    df["false_negative"] = ""


# FOR EACH ROW, CHECK EACH CONDITION
    for i, row in df.iterrows():
        tp, tn, fp, fn = [], [], [], []

        for cond in conditions:
            orig = str(row.get(f"{cond}_original", "")).strip().lower()
            ai = str(row.get(f"{cond}_ai", "")).strip().lower()

            if orig == "positive" and ai == "positive":
                tp.append(cond)
            elif orig == "negative" and ai == "negative":
                tn.append(cond)
            elif orig == "negative" and ai == "positive":
                fp.append(cond)
            elif orig == "positive" and ai == "negative":
                fn.append(cond)

        df.at[i, "true_positive"] = ", ".join(tp)
        df.at[i, "true_negative"] = ", ".join(tn)
        df.at[i, "false_positive"] = ", ".join(fp)
        df.at[i, "false_negative"] = ", ".join(fn)

    df.to_excel(LABEL_OUTPUT, index=False)
    print(f" Labeled predictions saved to {LABEL_OUTPUT}")
    return df


# CONFUSION MATRIX
# Normalize column names (lowercase, underscores):
def build_confusion(df):
    norm_df = df.copy()
    norm_df.columns = (
        norm_df.columns.str.strip()
        .str.lower()
        .str.replace(" ", "_")
    )

    result_rows = []
# Count occurrences per condition:
    for cond in CLASSIFIERS:
        tp = norm_df["true_positive"].str.contains(
            cond, case=False, na=False
        ).sum()
        fp = norm_df["false_positive"].str.contains(
            cond, case=False, na=False
        ).sum()
        tn = norm_df["true_negative"].str.contains(
            cond, case=False, na=False
        ).sum()
        fn = norm_df["false_negative"].str.contains(
            cond, case=False, na=False
        ).sum()

        result_rows.append([cond, tp, fp, tn, fn, "", "", "", "", "", ""])

    cm_df = pd.DataFrame(result_rows, columns=OUTPUT_COLUMNS)

    # Write final Excel with formulas
    wb = Workbook()
    ws = wb.active
    ws.title = "Confusion Matrix"

    for r in dataframe_to_rows(cm_df, index=False, header=True):
        ws.append(r)

# Excel formulas for Sensitivity (F) and Specificity (G) and other columns:
    for i in range(2, len(cm_df) + 2):
        ws[f"F{i}"] = f"=IFERROR(B{i}/(B{i}+E{i}),0)"  # Sensitivity
        ws[f"G{i}"] = f"=IFERROR(D{i}/(D{i}+C{i}),0)"  # Specificity
        # Format Sensitivity and Specificity columns as percentage
        ws[f"F{i}"].number_format = "0.00%"
        ws[f"G{i}"].number_format = "0.00%"
        # Check column formula: =SUM(true_Positive+false_Positive)
        ws[f"H{i}"] = (
            f"=SUM(B{i}+C{i})"
        )
        # Positive Ground Truth formula: =SUM(B2:E2)
        ws[f"I{i}"] = f"=SUM(B{i}:E{i})"
        # Negative Ground Truth formula: =SUM(D2:C2)
        ws[f"J{i}"] = f"=SUM(D{i}:C{i})"
        # Ground Truth Check formula: =SUM(I2:J2)
        ws[f"K{i}"] = f"=SUM(I{i}:J{i})"

# Creates a new Excel sheet named "Input Data
    ws_input = wb.create_sheet(title="Input Data")
# DataFrame into row format that Excel understands
    for r in dataframe_to_rows(
        df,
        index=False,
        header=True
    ):
        ws_input.append(r)
# Center-align cells in both sheets
    for sheet in [ws, ws_input]:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )
#  Saves the final Excel file
    wb.save(CONFUSION_OUTPUT)
    print(f" Confusion matrix saved to {CONFUSION_OUTPUT}")


#  MAIN Run AND Loads input Excel into DataFrame
# check for input file existence AND print completion message
def main():
    if not Path(INPUT_FILE).exists():
        print(f"Input not found: {INPUT_FILE}")
        return

    df = pd.read_excel(INPUT_FILE)
    df = label_predictions(df)
    build_confusion(df)
    print("\n All done — pipeline completed successfully!")


# run script directly
if __name__ == "__main__":
    main()
